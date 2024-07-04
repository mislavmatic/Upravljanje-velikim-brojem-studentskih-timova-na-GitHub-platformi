from openpyxl import load_workbook
from sqlalchemy import create_engine, ForeignKey, Column, BigInteger, VARCHAR
from sqlalchemy.orm import declarative_base, sessionmaker, relationship

Base = declarative_base()

class Asistent(Base):
    __tablename__  = "Asistenti"

    as_id = Column("as_id", BigInteger, primary_key=True)
    name_asist = Column("name_asist",  VARCHAR(70))

    demos = relationship('Demos', back_populates='asistent')

    def __init__(self, name) -> None:
        self.name_asist = name
    
    def __repr__(self):
        return f"({self.as_id}) {self.name_asist}"

class Demos(Base):
    __tablename__  = "Demosi"

    de_id = Column("de_id", BigInteger, primary_key=True)
    name_demos = Column("name_demos",  VARCHAR(70))
    as_id = Column(BigInteger, ForeignKey("Asistenti.as_id"))

    asistent = relationship('Asistent', back_populates='demos')
    groups = relationship('Grupa', back_populates='demos')

    def __init__(self, name, as_id) -> None:
        self.name_demos = name
        self.as_id = as_id                  # ~_~

    def __repr__(self) -> str:
        return f"{self.name_demos} kod asistenta {self.as_id}"
    
class Student(Base):
    __tablename__  = "Studenti"

    s_email = Column("s_email", VARCHAR(70), primary_key=True)
    name_stud = Column("name_stud",  VARCHAR(70))

    stud_grupe = relationship('StudGrupa', back_populates='student')

    def __init__(self, email, name) -> None:
        self.s_email = email
        self.name_stud = name

    def __repr__(self) -> str:
        return f"{self.name_stud}"


class Zadatak(Base):
    __tablename__  = "Zadaci"

    name_zad = Column("name_zad",  VARCHAR(70), primary_key=True)

    groups = relationship('Grupa', back_populates='zadatak')

    def __init__(self, name) -> None:
        self.name_zad = name

class Grupa(Base):
    __tablename__  = "Grupe"

    github = Column("github", VARCHAR(70), primary_key=True)
    name_g = Column("name_g",  VARCHAR(70))
    name_zad = Column(VARCHAR(70), ForeignKey("Zadaci.name_zad"))
    de_id = Column(BigInteger, ForeignKey("Demosi.de_id"))
    leader_email = Column(VARCHAR(70), ForeignKey("Studenti.s_email"))

    zadatak = relationship('Zadatak', back_populates='groups')
    demos = relationship('Demos', back_populates='groups')
    students = relationship('StudGrupa', back_populates='group')

    def __init__(self, name_g, name_zad, demos_id, leader_email, github) -> None:
        self.name_g = name_g
        self.name_zad = name_zad
        self.de_id = demos_id
        self.leader_email = leader_email
        self.github = github

class StudGrupa(Base):
    __tablename__  = "StudGrupa"

    s_email = Column(VARCHAR(70), ForeignKey("Studenti.s_email"), primary_key=True)
    github = Column(VARCHAR(70), ForeignKey("Grupe.github"), primary_key=True)

    student = relationship('Student', back_populates='stud_grupe')
    group = relationship('Grupa', back_populates='students')

    def __init__(self, email, github) -> None:
        self.s_email = email
        self.github = github


# class StudOcjena(Base):
#     __tablename__ = "StudOcjena"

def create_session():
        engine = create_engine("postgresql+psycopg2://postgres:bazepodataka@localhost:5434/ZavrsniRad", echo=True) #@localhost:5432/zavrsnirad
        Base.metadata.create_all(bind=engine)

        Session = sessionmaker(bind=engine)
        session = Session()
        return session

def parsiranje(file):
    session =  create_session()
    

    wb = load_workbook(file)

    for sheet in wb.worksheets:
        asistent_nehardkodirani = Asistent(sheet.title)
        session.add(asistent_nehardkodirani)
        session.commit()

        work_sheet = sheet

        max_row = work_sheet.max_row

        counter = 0
        for cell in work_sheet['B']:
            if cell.row == counter or cell.value == "KRAJ":
                break

            if cell.value == "E-mail":       #AKO POSTOJI GRUPA KOJA SE ZOVE E-mail ONDA OVO BAŠ I NEĆE RADITi
                #cell.row je red u kojem se nalazi ćelija E-mail
                group_name_cell = 'B' + str(cell.row - 1)
                prez_st_cell = 'C' + str(cell.row)
                ime_st_cell = 'D' + str(cell.row)
                zadatak_name_cell = 'G' + str(cell.row) # 'Zadatak:' je u ćeliji (F, cell.row), a ime zadatka u (G, cell.row)
                demos_name_cell = 'G' + str(cell.row + 1)
                github_link_cell = 'G' + str(cell.row + 2)

                asist_id = asistent_nehardkodirani.as_id
                demos = Demos(work_sheet[demos_name_cell].value, asist_id) 
                session.add(demos)
                session.commit()

                kursor = cell.row + 1
                cell_stud = work_sheet['C' + str(kursor)].value
                while cell_stud != "Prezime" and cell_stud != '\\' and kursor <= max_row:   # jel potreban ovaj prvi != Prezime ?
                    if work_sheet['C' + str(kursor)].value == "\\":
                        break
                    cell_stud = work_sheet['C' + str(kursor)].value
                    stud_ime_prez = work_sheet['D' + str(kursor)].value + " " + cell_stud
                    stud_email = work_sheet['B' + str(kursor)].value
                    kursor += 1
                    stud = Student(stud_email, stud_ime_prez)
                    session.add(stud)
                    session.commit()

                zad_name = work_sheet[zadatak_name_cell].value
                zadatak = Zadatak(zad_name)
                zadatak_old = session.query(Zadatak).filter_by(name_zad=zad_name).first()
                if zadatak_old is None:
                    session.add(zadatak)
                    session.commit()
                

                github_link = work_sheet[github_link_cell].value

                group_name = work_sheet[group_name_cell].value

                demos_id = session.query(Demos).filter_by(name_demos=work_sheet[demos_name_cell].value).first().de_id
    
                leader_email = work_sheet['B' + str(cell.row + 1)].value

                grupa = Grupa(group_name, zad_name, demos_id, leader_email, github_link)
                session.add(grupa)
                session.commit()

                kursor = cell.row + 1
                cell_stud = work_sheet['C' + str(kursor)].value
                while cell_stud != "Prezime" and cell_stud != '\\' and kursor <= max_row:
                    if work_sheet['C' + str(kursor)].value == "\\":
                        break
                    stud_email = work_sheet['B' + str(kursor)].value
                    kursor += 1
                    studGrup = StudGrupa(stud_email, github_link)
                    session.add(studGrup)
                    session.commit()
                    

            counter += 1

    session.close()
