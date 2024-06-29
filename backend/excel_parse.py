from openpyxl import Workbook, load_workbook

from sqlalchemy import create_engine, ForeignKey, Column, String, Integer, CHAR, BigInteger, VARCHAR
from sqlalchemy.orm import declarative_base, sessionmaker, relationship

Base = declarative_base()

class Asistent(Base):
    __tablename__  = "Asistenti"

    as_id = Column("as_id", BigInteger, primary_key=True)
    name_asist = Column("name_asist",  VARCHAR(70))

    demos = relationship('Demos', back_populates='asistent')

    def __init__(self, name) -> None:
        self.name_asist = name
        #self.last_name_asist = last
    
    def __repr__(self):
        return f"({self.as_id}) {self.name_asist}"

class Demos(Base):
    __tablename__  = "Demosi"

    de_id = Column("de_id", BigInteger, primary_key=True)
    name_demos = Column("name_demos",  VARCHAR(70))
    as_id = Column(BigInteger, ForeignKey("Asistenti.as_id"))

    asistent = relationship('Asistent', back_populates='demos')
    groups = relationship('Grupa', back_populates='demos')

    def __init__(self, name, as_id) -> None:
        self.name_demos = name
        #self.last_name_demos = last
        self.as_id = as_id                  # ~_~

    def __repr__(self) -> str:
        return f"{self.name_demos} kod asistenta {self.as_id}"
    
class Student(Base):
    __tablename__  = "Studenti"

    s_id = Column("s_id", BigInteger, primary_key=True)
    name_stud = Column("name_stud",  VARCHAR(70))

    stud_grupe = relationship('StudGrupa', back_populates='student')

    def __init__(self, name) -> None:
        self.name_stud = name
        #self.last_name_demos = last
        #JMBAG?

    def __repr__(self) -> str:
        return f"{self.name_stud}"


class Zadatak(Base):
    __tablename__  = "Zadaci"

    name_zad = Column("name_zad",  VARCHAR(70), primary_key=True)

    groups = relationship('Grupa', back_populates='zadatak')

    def __init__(self, name) -> None:
        self.name_zad = name

class Grupa(Base):
    __tablename__  = "Grupe"

    github = Column("github", VARCHAR(70), primary_key=True)
    name_g = Column("name_g",  VARCHAR(70))
    name_zad = Column(VARCHAR(70), ForeignKey("Zadaci.name_zad"))
    de_id = Column(BigInteger, ForeignKey("Demosi.de_id"))
    leader_id = Column(BigInteger, ForeignKey("Studenti.s_id"))

    zadatak = relationship('Zadatak', back_populates='groups')
    demos = relationship('Demos', back_populates='groups')
    students = relationship('StudGrupa', back_populates='group')

    def __init__(self, name_g, name_zad, demos_id, leader_id, github) -> None:
        self.name_g = name_g
        self.name_zad = name_zad
        self.de_id = demos_id
        self.leader_id = leader_id
        self.github = github

class StudGrupa(Base):
    __tablename__  = "StudGrupa"

    # s_id = Column("s_id", BigInteger, primary_key=True)
    # github = Column(VARCHAR(70), ForeignKey("Grupe.github"))
    s_id = Column(BigInteger, ForeignKey("Studenti.s_id"), primary_key=True)
    github = Column(VARCHAR(70), ForeignKey("Grupe.github"), primary_key=True)

    student = relationship('Student', back_populates='stud_grupe')
    group = relationship('Grupa', back_populates='students')

    def __init__(self, s_id, github) -> None:
        self.s_id = s_id
        self.github = github

    # def __repr__(self) -> str:
    #     return f"{self.name_stud}"

def create_session():
        engine = create_engine("postgresql+psycopg2://postgres:bazepodataka@localhost:5434/ZavrsniRad", echo=True) #@localhost:5432/zavrsnirad
        Base.metadata.create_all(bind=engine)

        Session = sessionmaker(bind=engine)
        session = Session()
        return session

def parsiranje(file):
    session =  create_session()
    

    wb = load_workbook(file)

    for sheet in wb.worksheets:
        # sheet je asistent pa nek se splita po razmaku i doda odma u bazu
        # onda se demosi trebaju dodati (treba poznavati as_id)

        #asistent_hardkodirani = Asistent('AJ_03_04')
        asistent_hardkodirani = Asistent(sheet.title)
        session.add(asistent_hardkodirani)
        session.commit()

        #raise NotImplementedError("Bleh")

        # active_sheet = wb.active
        # print(active_sheet)

        #trenutno ćemo hardkodirat da se samo sheet 'AJ_03_04' koristi
        #work_sheet = wb['AJ_03_04']
        work_sheet = sheet
        #print(work_sheet['B2'])
        #print(work_sheet['B2'].value)
        max_row = work_sheet.max_row

        #kursor = work_sheet['B2']
        counter = 0
        for cell in work_sheet['B']:
            if cell.row == counter or cell.value == "KRAJ":
                break

            if cell.value == "JMBAG":       #AKO POSTOJI GRUPA KOJA SE ZOVE JMBAG ONDA OVO BAŠ I NEĆE RADITI
                #print(cell.row, end=' ')
                #cell.row je red u kojem se nalazi ćelija JMBAG
                group_name_cell = 'B' + str(cell.row - 1)
                prez_st_cell = 'C' + str(cell.row)
                ime_st_cell = 'D' + str(cell.row)
                zadatak_name_cell = 'G' + str(cell.row) # 'Zadatak:' je u ćeliji (F, cell.row), a ime zadatka u (G, cell.row)
                demos_name_cell = 'G' + str(cell.row + 1)
                github_link_cell = 'G' + str(cell.row + 2)

                #sad treba parsirat onim redoslijedom i 
                asist_id = asistent_hardkodirani.as_id
                demos = Demos(work_sheet[demos_name_cell].value, asist_id) 
                #sad treba demosa nadodat u bazu (jedinstvenost???)
                session.add(demos)
                session.commit()

                #raise NotImplementedError("Demos?")

                kursor = cell.row + 1
                cell_stud = work_sheet['C' + str(kursor)].value
                while cell_stud != "Prezime" and cell_stud != '\\' and kursor <= max_row:
                    #print(cell_stud)
                    if work_sheet['C' + str(kursor)].value == "\\":
                        break
                    cell_stud = work_sheet['C' + str(kursor)].value
                    stud_ime_prez = work_sheet['D' + str(kursor)].value + " " + cell_stud
                    kursor += 1
                    #print(stud_ime_prez)
                    stud = Student(stud_ime_prez)
                    session.add(stud)
                    session.commit()
                    #dodaj u bazu studenta

                #raise NotImplementedError("Prvi studenti?")
                zad_name = work_sheet[zadatak_name_cell].value
                #print(zad_name)
                zadatak = Zadatak(zad_name)
                #dodaj u bazu (jedinstvenost?)
                zadatak_old = session.query(Zadatak).filter_by(name_zad=zad_name).first()
                if zadatak_old is None:
                    session.add(zadatak)
                    session.commit()
                

                #raise NotImplementedError("Zadaci?")

                github_link = work_sheet[github_link_cell].value
                #print(github_link)

                group_name = work_sheet[group_name_cell].value
                #print(group_name)

                #TREBA IZVUĆI DEMOS_ID i leader_id IZ BAZE
                demos_id = session.query(Demos).filter_by(name_demos=work_sheet[demos_name_cell].value).first().de_id
                #raise NotImplementedError("demos_id?")
                #DEMOS_ID = 1 #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                #leader_id = 1 #!!!!!!!!!!!!!!¨!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                leader_name = work_sheet['D' + str(cell.row + 1)].value + " " + work_sheet['C' + str(cell.row + 1)].value     #pretpostvaka je da je leader prvi od studenata u grupi
                
                leader_id = session.query(Student).filter_by(name_stud=leader_name).first().s_id
                #print(leader_id)
                #raise NotImplementedError("id")
                grupa = Grupa(group_name, zad_name, demos_id, leader_id, github_link)
                #dodaj u bazu grupu
                session.add(grupa)
                session.commit()

                #raise NotImplementedError("grupa?")
                kursor = cell.row + 1
                cell_stud = work_sheet['C' + str(kursor)].value
                while cell_stud != "Prezime" and cell_stud != '\\' and kursor <= max_row:
                    #print(cell_stud)
                    if work_sheet['C' + str(kursor)].value == "\\":
                        break
                    cell_stud = work_sheet['C' + str(kursor)].value
                    stud_ime_prez = work_sheet['D' + str(kursor)].value + " " + cell_stud
                    kursor += 1
                    #print(stud_ime_prez)
                    s_id = session.query(Student).filter_by(name_stud=stud_ime_prez).first().s_id
                    studGrup = StudGrupa(s_id, github_link)
                    session.add(studGrup)
                    session.commit()
                    
                #print("")

            counter += 1

    session.close()     # ?
